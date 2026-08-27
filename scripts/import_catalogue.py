#!/usr/bin/env python3
"""
استيراد كتالوج المركبات من ملف Excel.

    python3 scripts/import_catalogue.py db/sources/morocco-cars-2026.xlsx

كيخرّج جوج حوايج من نفس المصدر باش مايتفرّقوش:

  src/lib/data/catalog.ts  — نسخة مرفقة مع الموقع. المتصفح محتاجها
    فالحين للاقتراحات وتحليل الدارجة، ماعندوش وقت يسنّى طلب
    للخادم فكل حرف كيكتبو المستخدم.

  db/catalog.sql           — نفس المعطيات لقاعدة البيانات، باش
    تقدر تزيد ولا تبدّل ماركة بلا ما تعاود تنشر الموقع.

الدراجات النارية ماكايناش فهاد الملف (كتالوج سيارات)، فكنحتافظو
بالموجود. ونفس الشي للفئات (berline, suv…): الملف ماكيعطيهاش،
فكناخدوها من النسخة القديمة ملي يكون الموديل معروف.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

# أسماء الماركات كيفرقو بين المصادر — كنوحّدوهم على اللي كنستعملو
# فالموقع باش الروابط والشعارات يبقاو خدّامين.
CANON = {
    "mercedes-benz": "Mercedes",
    "seat": "Seat",
    "škoda": "Skoda",
    "skoda": "Skoda",
    "mini": "Mini",
    "smart": "Smart",
}

ROOT = Path(__file__).resolve().parent.parent
CATALOG_TS = ROOT / "src/lib/data/catalog.ts"
CATALOG_SQL = ROOT / "db/catalog.sql"


def canon(name: str) -> str:
    n = name.strip()
    return CANON.get(n.lower(), n)


def clean(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s != "Not specified" else None


def read_existing() -> dict[str, dict]:
    """الموديلات اللي عندنا دابا — فيها الفئات والدراجات."""
    text = CATALOG_TS.read_text(encoding="utf-8")
    out: dict[str, dict] = {}
    for line in text.splitlines():
        m = re.match(r'\s*\{ kind: "(\w+)", make: "([^"]+)", model: "([^"]+)"(.*)\},', line)
        if not m:
            continue
        kind, make, model, rest = m.groups()
        item = {"kind": kind, "make": make, "model": model}
        for key in ("body",):
            k = re.search(rf'{key}: "([^"]+)"', rest)
            if k:
                item[key] = k.group(1)
        for key in ("cv", "cc", "doors"):
            k = re.search(rf"{key}: (\d+)", rest)
            if k:
                item[key] = int(k.group(1))
        out[f'{kind}|{make}|{model}'] = item
    return out


def ts_line(c: dict) -> str:
    parts = [f'kind: "{c["kind"]}"', f'make: {json.dumps(c["make"], ensure_ascii=False)}',
             f'model: {json.dumps(c["model"], ensure_ascii=False)}']
    if c.get("body"):
        parts.append(f'body: "{c["body"]}"')
    if c.get("cc"):
        parts.append(f'cc: {c["cc"]}')
    elif c.get("cv"):
        parts.append(f'cv: {c["cv"]}')
    if c.get("doors"):
        parts.append(f'doors: {c["doors"]}')
    return "  { " + ", ".join(parts) + " },"


def sq(v: str | None) -> str:
    return "NULL" if v is None else "'" + v.replace("'", "''") + "'"


def main() -> None:
    if len(sys.argv) < 2:
        sys.exit("الاستعمال: python3 scripts/import_catalogue.py <ملف.xlsx>")
    src = Path(sys.argv[1])
    wb = openpyxl.load_workbook(src, data_only=True)

    # ---- الماركات ----
    brands: dict[str, dict] = {}
    for i, row in enumerate(wb["Brands"].iter_rows(values_only=True)):
        if i == 0 or not row[0]:
            continue
        make = canon(str(row[0]))
        brands[make] = {"make": make, "country": clean(row[1]), "group": clean(row[2])}

    # ---- الموديلات ----
    existing = read_existing()
    models: dict[str, dict] = {k: v for k, v in existing.items() if v["kind"] == "moto"}

    for i, row in enumerate(wb["Models"].iter_rows(values_only=True)):
        if i == 0 or not row[0] or not row[1]:
            continue
        make, model = canon(str(row[0])), str(row[1]).strip()
        key = f"car|{make}|{model}"
        old = existing.get(key, {})
        item = {"kind": "car", "make": make, "model": model}
        for f in ("body", "cv", "doors"):
            if old.get(f):
                item[f] = old[f]
        models[key] = item

    # اللي كان عندنا وماكاينش فالملف كيبقى — عمرنا ما نحيّدو موديل
    for key, item in existing.items():
        models.setdefault(key, item)

    ordered = sorted(models.values(), key=lambda c: (c["kind"], c["make"], c["model"]))

    # ---- الملف المرفق ----
    text = CATALOG_TS.read_text(encoding="utf-8")
    head = text[: text.index("export const CATALOG")]
    tail = text[text.index("];\n") + 3:]
    body = "\n".join(ts_line(c) for c in ordered)
    CATALOG_TS.write_text(f"{head}export const CATALOG: CatalogItem[] = [\n{body}\n];\n{tail}", encoding="utf-8")

    # ---- SQL ----
    brand_rows = [f"  ({sq(b['make'])}, 'car', {sq(b['country'])}, {sq(b['group'])})"
                  for b in sorted(brands.values(), key=lambda b: b["make"])]
    moto_makes = sorted({c["make"] for c in ordered if c["kind"] == "moto"})
    brand_rows += [f"  ({sq(m)}, 'moto', NULL, NULL)" for m in moto_makes]

    model_rows = [f"  ('{c['kind']}', {sq(c['make'])}, {sq(c['model'])}, "
                  f"{sq(c.get('body'))})" for c in ordered]

    brands_sql = ",\n".join(brand_rows)
    models_sql = ",\n".join(model_rows)

    CATALOG_SQL.write_text(f"""-- ============================================================
-- كتالوج المركبات — ماركات وموديلات كتّباع فالمغرب
--
-- مولّد من: {src.name}
-- المصدر: Moteur.ma / Ovoiture (كتالوج السيارات الجديدة 2026)
--
-- مرجع فقط: ماكاين لا أثمنة لا كيلومتراج لا بائعين. الإعلانات
-- الحقيقية كتجي من الناس.
--
-- كيتعاود تشغيلو بلا مشكل: ON CONFLICT كيحيّن اللي كاين.
-- الاستعمال: Neon SQL Editor → الصق → Run
-- ============================================================

INSERT INTO catalog_brands (make, kind, country, parent_group) VALUES
{brands_sql}
ON CONFLICT (make, kind) DO UPDATE
  SET country = EXCLUDED.country, parent_group = EXCLUDED.parent_group;

INSERT INTO catalog_models (kind, make, model, body) VALUES
{models_sql}
ON CONFLICT (kind, make, model) DO UPDATE
  SET body = COALESCE(EXCLUDED.body, catalog_models.body);

SELECT kind, count(*) AS n FROM catalog_models GROUP BY kind ORDER BY kind;
""", encoding="utf-8")

    cars = sum(1 for c in ordered if c["kind"] == "car")
    motos = len(ordered) - cars
    print(f"✓ {len(brands)} ماركة · {cars} سيارة · {motos} دراجة")
    print(f"  → {CATALOG_TS.relative_to(ROOT)}")
    print(f"  → {CATALOG_SQL.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
